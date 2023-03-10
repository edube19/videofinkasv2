from openpyxl import Workbook
from recursos.re_excel import *
from openpyxl.styles import *

def generar_excel(datos_dpto_estacionamiento,datos_subsecciones,datos_finca,finca,cantidad_propietarios,fecha_emision,fecha_vencimiento,tipo,tipo_moneda='S/.',mensaje_extra='Mensaje extra al pie de pagina',n_excel=1):
    #para leer el json
    solo_dpto = False
    #DATOS DE LA FINCA
    info_finca = datos_finca[0]
    direccion = info_finca["Direccion"]
    Nombre = info_finca["Nombre"]

    #DATOS SUBSECCION
    info_subseccion = datos_subsecciones[0]
    mes = info_subseccion["Mes"]
    año = info_subseccion["Year"]

    w,h = 4,cantidad_propietarios
    lista_general = [[0 for x in range(w)] for y in range(h)]#creando la matriz general
    json_rutas_excel = []
    for i in range(cantidad_propietarios):
        dato_dpto_estc = datos_dpto_estacionamiento[i]
        propietario=dato_dpto_estc["Nombres_y_Apellidos"]
        id_departamento = dato_dpto_estc["Departamentos"][0]["ID_Departamentos"]
        porc_participacion = dato_dpto_estc["Departamentos"][0]["Porcentaje_Participacion"]
        num_estacionamiento = dato_dpto_estc["Estacionamientos"][0]["Numero_Estacionamiento"]

        if tipo==1:
            #if num_estacionamiento == "" and id_departamento != "":
            if num_estacionamiento == "":
               print('SOLO DEPARTAMENTO\n')
               solo_dpto = True
            else:
                print('TANTO DPTO COMO ESTACIONAMIENTO\n') 
                solo_dpto = False
        else:
            solo_dpto = True
            print('SOLO ESTACIONAMIENTO\n')

        estado = 0

        try:
            #CREANDO EL EXCEL
            total_monto=0
            book=Workbook()
            sheet= book.active
            
            #[17,18,14,10,17] original
            ancho_columnas_parametros(sheet)
            combinar_celdas(sheet,'B1','D1','JUNTA DE PROPIETARIOS')
            sheet['B1'].alignment=Alignment(horizontal="center")
            combinar_celdas(sheet,'B2','D2',Nombre)
            sheet['B2'].alignment=Alignment(horizontal="center")
            combinar_celdas(sheet,'B3','D3',direccion)
            sheet['B3'].alignment=Alignment(horizontal="center")
            combinar_celdas(sheet,'B4','D4','RECIBO POR CUOTA DE MANTENIMIENTO')
            sheet['B4'].alignment=Alignment(horizontal="center")
            formato_celdas(sheet,'B1','Arial',9,True,True,'000000',True)
            combinar_celdas(sheet,'A1','A4')
            combinar_celdas(sheet,'E1','E4')

            if tipo == 1:#
                sheet['A5']='Departamento:'
                sheet['B5']=id_departamento
                sheet['B5'].alignment=Alignment(horizontal='center')
            else:
                sheet['A5']='Estacionamiento:'
                sheet['B5']=num_estacionamiento
                sheet['B5'].alignment=Alignment(horizontal='center')
            
            sheet['A6']='Propietario:'
            sheet['A7']='Periodo:'
            #mes = 'SETIEMBRE'
            #año = '2022'
            periodo = nombre_mes(mes,año)#poner el mes segun el numero de este 
            sheet['B7']=periodo
            sheet['B7'].alignment=Alignment(horizontal='center')

            if tipo == 1 and solo_dpto == False:#tiene ambos
                sheet['C5']='Estacionamiento:'
                sheet.merge_cells('D5:E5')
                sheet['D5']=num_estacionamiento
                sheet['D5'].alignment=Alignment(horizontal='center')
            elif solo_dpto:
                sheet.merge_cells('B5:E5')

            sheet.merge_cells('B6:E6')

            sheet['B6']=propietario
            sheet['B6'].alignment=Alignment(horizontal='center')
            sheet['C7']='Total Porc. Part.'

            if solo_dpto ==False:
                sheet.merge_cells('D5:E5')
                sheet['D5']=num_estacionamiento
                sheet['D5'].alignment=Alignment(horizontal='center')

            sheet.merge_cells('D7:E7')
            sheet['D7']=porc_participacion
            sheet['D7'].alignment=Alignment(horizontal='center')

            sheet['D8']='Total'
            formato_celdas(sheet,'D8','Calibri',11,False,True,'000000',True)
            sheet['D8'].alignment=Alignment(horizontal='center')

            sheet['E8']='Importe'
            formato_celdas(sheet,'E8','Calibri',11,False,True,'000000',True)
            sheet['E8'].alignment=Alignment(horizontal='center')
            #----------------------------------------------------------------
            #RECORRIENDO LOS DATOS DE CADA FINCA
            iterable=9
            #en esta seccion se llenan los datos de la BD
            #------------------SUBSECCIONES----------------------------------------------
            #plantillas=json[0]['Plantillas']
            secciones = info_subseccion["Seccion"]
            #print('CANT SECCIONES >>> ',str(len(secciones)))

            for j in range(len(secciones)):
                
                subseccion = secciones[j]['Subsecciones']
                nombre_seccion = secciones[j]['nombre']
                #print('CANT SUBSECCIONES >>> ',str(len(subseccion)))
                
                for s in range(len(subseccion)):
                    #secciones
                    #id_seccion = secciones[s]['ID_Seccion']
                    #print('NOMBRE DE LA SECCION >>> ',str(nombre_seccion))
                    #nombre = secciones[s]['nombre']
                    celda_seccion=f'A{iterable}'
                    sheet[celda_seccion]=nombre_seccion
                    
                    formato_celdas(sheet,celda_seccion,'Calibri',11,True,True,'000000',True)
                    iterable=iterable+1

                    nombre_subseccion = subseccion[s]['nombre']
                    
                    monto_subseccion = subseccion[s]['monto']
                    
                    descripcion_subseccion = subseccion[s]['descripcion']
                    
                    #subsecciones
                    celda_col1=f'A{iterable}'
                    valor_col1=nombre_subseccion
                    sheet[celda_col1] = valor_col1

                    #descripcion
                    celda_col2=f'B{iterable}'
                    celda_col2_fin = f'C{iterable}'
                    valor_col2=descripcion_subseccion
                    sheet[celda_col2] = valor_col2
                    combinar_celdas(sheet,celda_col2,celda_col2_fin)
                    sheet[celda_col2].alignment=Alignment(horizontal='center')

                    #monto
                    celda_col4=f'D{iterable}'
           
                    monto_float = "{:.2f}".format(monto_subseccion)
                 
                    if (tipo_moneda!='€'):
                        valor_col4=tipo_moneda +' '+str(monto_float)
    
                    else:
                        valor_col4=str(monto_float)+' '+tipo_moneda 
       
                    sheet[celda_col4] = valor_col4
                    sheet[celda_col4].alignment=Alignment(horizontal='center')

                    celda_col5=f'E{iterable}'        
                    
                    porc_participacion_float = float(porc_participacion)
                    
                    total_monto_fila=monto_subseccion*(porc_participacion_float/100)#ACA FALLA POR ALGUN MOTIVO ....
                    

                    monto_fila = monto_subseccion + total_monto_fila

                    total_monto=total_monto+monto_fila
                    total_monto_fila_float="{:.2f}".format(total_monto_fila)
                    
                    if (tipo_moneda!='€'):
                        valor_col5=tipo_moneda +' '+str(total_monto_fila_float)
                    else:
                        valor_col5=str(total_monto_fila_float)+' '+tipo_moneda

                    sheet[celda_col5] = valor_col5
                    sheet[celda_col5].alignment=Alignment(horizontal='center')

                    celda_final_suma=celda_col5
                    
                    iterable = iterable + 1
                    """for ss in range(len(subseccion)):
                        id_subseccion = subsecciones[ss]["ID_Subseccion"]
                        #nombre
                        nombre = subsecciones[ss]['nombre']

                        nombre_subseccion = id_subseccion+'-'+nombre
                        celda_col1=f'A{iterable}'
                        valor_col1=nombre_subseccion
                        sheet[celda_col1] = valor_col1

                        #descripcion
                        descripcion = subsecciones[ss]['descripcion']
                        celda_col2=f'B{iterable}'
                        celda_col2_fin = f'C{iterable}'
                        valor_col2=descripcion
                        sheet[celda_col2] = valor_col2
                        combinar_celdas(sheet,celda_col2,celda_col2_fin)
                        sheet[celda_col2].alignment=Alignment(horizontal='center')
                        #monto
                        monto = subsecciones[ss]['monto']
                        celda_col4=f'D{iterable}'
                        monto_float = "{:.2f}".format(monto)
                        if (tipo_moneda!='€'):
                            valor_col4=tipo_moneda +' '+str(monto_float)
                        else:
                            valor_col4=str(monto_float)+' '+tipo_moneda 
                        sheet[celda_col4] = valor_col4
                        sheet[celda_col4].alignment=Alignment(horizontal='center')

                        celda_col5=f'E{iterable}'        

                        total_monto_fila=monto*(porc_participacion/100)

                        monto_fila = monto + total_monto_fila

                        total_monto=total_monto+monto_fila
                        total_monto_fila_float="{:.2f}".format(total_monto_fila)
                        if (tipo_moneda!='€'):
                            valor_col5=tipo_moneda +' '+str(total_monto_fila_float)
                        else:
                            valor_col5=str(total_monto_fila_float)+' '+tipo_moneda

                        sheet[celda_col5] = valor_col5
                        sheet[celda_col5].alignment=Alignment(horizontal='center')

                        celda_final_suma=celda_col5
                        iterable = iterable + 1"""
            
            #La suma total
            total_monto_float="{:.2f}".format(total_monto)
            
            if (tipo_moneda!='€'):
                celda_suma_expresion=tipo_moneda +' '+str(total_monto_float)
            else:
                celda_suma_expresion=str(total_monto_float)+' '+tipo_moneda
            
            celda_total=f'A{iterable}'
            celda_total_valor=f'E{iterable}'
            sheet[celda_total] = 'TOTAL'

            formato_celdas(sheet,celda_total,'Calibri',11,True,False,'000000')
            sheet[celda_total_valor]=celda_suma_expresion
            sheet[celda_total_valor].alignment=Alignment(horizontal='center')
            formato_celdas(sheet,celda_total_valor,'Calibri',11,True,False,'000000')

            bordear_celdasv1(sheet,'A1','F8')

            iterable= iterable + 1

            #Fechas emision
            celda_fecha_emision=f'A{iterable}'
            sheet[celda_fecha_emision]='Fecha de emision' #inicial
            
            celda_fecha_emision=f'A{iterable+1}'
            #valor_fecha_emision='01/09/2022'
            sheet[celda_fecha_emision]=fecha_emision
            sheet[celda_fecha_emision].alignment=Alignment(horizontal='center')

            #fecha vencimiento
            celda_fecha_vencimento=f'B{iterable}'
            sheet[celda_fecha_vencimento]='Fecha de vencimiento'

            celda_fecha_vencimento=f'B{iterable+1}'
            #valor_fecha_vencimento='07/07/2022'
            sheet[celda_fecha_vencimento]=fecha_vencimiento
            sheet[celda_fecha_vencimento].alignment=Alignment(horizontal='center')

            #N° de cuenta
            celda_ncuenta=f'C{iterable}'
            sheet[celda_ncuenta]='N° Cuenta'

            celda_CCI=f'C{iterable+1}'
            sheet[celda_CCI]='CCI'
            sheet[celda_CCI].alignment=Alignment(horizontal='center')

            valor_ini_ncuenta=f'D{iterable}'
            ncuenta='194-123456789'
            sheet[valor_ini_ncuenta]=ncuenta
            
            valor_fin_ncuenta=f'E{iterable}'
            #CCI
            valor_ini_CCI=f'D{iterable+1}'
            cci="0021194132456789" #final
            sheet[valor_ini_CCI]=cci

            valor_fin_CCI=f'E{iterable+1}'
            ultima_fila=sheet.max_row
            bordear_celdasv2(sheet,iterable,ultima_fila-1,True)
            combinar_celdas(sheet,valor_ini_ncuenta,valor_fin_ncuenta)
            combinar_celdas(sheet,valor_ini_CCI,valor_fin_CCI)
            sheet[valor_ini_CCI].alignment=Alignment(horizontal='center')

            for fila in range(8,iterable):
                for columna in range(1,6):
                    if (columna==1):
                        sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(border_style=None, color='000000'),
                                                                top=Side(border_style=None, color='000000'),
                                                                bottom=Side(border_style=None, color='000000'))
                    if (fila==8):
                        sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                                right=Side(border_style=None, color='000000'),
                                                                top=Side(border_style='thin', color='000000'),
                                                                bottom=Side(border_style=None, color='000000'))
                    if (columna==5):
                        bordear_lado('derecha',sheet,fila,columna)
                        
                    if (fila==(iterable)):
                        bordear_lado('abajo',sheet,fila,columna)
            
            #fila del titular ultima_fila+1
            celda_vacia_ini=f'A{ultima_fila+1}'
            celda_vacia_fin=f'B{ultima_fila+1}'
            combinar_celdas(sheet,celda_vacia_ini,celda_vacia_fin)

            #ultimas celdas
            celda_titular_ini = f'C{ultima_fila+1}'
            celda_titular_fin = f'E{ultima_fila+1}'
            titular=propietario
            texto_titular = 'BCP - Titular: '+titular
            combinar_celdas(sheet,celda_titular_ini,celda_titular_fin,texto_titular)
            sheet[celda_titular_ini].alignment=Alignment(horizontal='center')

            #ultimas celdas de texto
            celdaini_mensaje =f'A{ultima_fila+2}'
            celdafin_mensaje =f'E{ultima_fila+3}'
            end = ultima_fila+3
            bordear_celdasv2(sheet,ultima_fila,end)
            
            combinar_celdas(sheet,celdaini_mensaje,celdafin_mensaje,mensaje_extra)
            sheet.cell(ultima_fila+2,1).alignment = Alignment(horizontal='center',vertical='center')
            poner_imagenes(sheet)
            
            sheet.cell(1,5).alignment = Alignment(horizontal='center',vertical='center')

            sheet.cell(8,1).border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style=None, color='000000'),
                                            top=Side(border_style=None, color='000000'),
                                            bottom=Side(border_style=None, color='000000'))

            nombres_completos = propietario.replace(' ','_')#habia un problema si dejaba los nombres con espacio
            nombre_excel = f'propietario_{nombres_completos}.xlsx'
            ruta_excel_real = guardar_ruta_excel(nombre_excel,finca,nombres_completos,año,mes)
            book.save(ruta_excel_real)
            agregar_fecha()
            #aumentando el numero del excel a guardar
            n_excel=n_excel+1
            
            #si se ve esto en pantalla, se genero el excel con exito
            print('Excel creado del propietario '+nombres_completos)
            
            #cerrar el libro
            book.close()  
            
            #guardando datos en la lista general
            lista_general[i][0] = estado #estado
            lista_general[i][1] = nombres_completos #cliente
            #lista_general[i][2] = excel_codificado #codificacion del excel del cliente
            diccionario_info = {"cliente":nombres_completos,"estado":estado,"ruta_excel":ruta_excel_real}
            json_rutas_excel.append(diccionario_info)
        except Exception as e:
            print('error → ', str(e))
            estado = 1 #fallo
            lista_general[i][0] = estado #estado
            lista_general[i][1] = propietario
            diccionario_info = {"cliente":nombres_completos,"estado":estado,"ruta_excel":'Error al procesar el documento: '+e}
            json_rutas_excel.append(diccionario_info)
    print('ACABO LA CREACION DEL EXCEL')
    return lista_general,json_rutas_excel,año,mes