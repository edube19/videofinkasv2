import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import xlwings as xw
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image
import pathlib
import pythoncom
import shutil
from datetime import datetime
import uuid 
import base64
#from libs.database import conexion,conexionv2
from bson import json_util
import json
import bcrypt
from conexion.conexion import conexion
from sqlalchemy import create_engine, MetaData, Table,select

def combinar_celdas(sheet,celda_inicial,celda_final,texto=''):
    sheet.merge_cells(f'{celda_inicial}:{celda_final}')
    if texto!='':
        sheet[celda_inicial]=texto

def formato_celdas(sheet,celda,fuente,tamaño,negrita=False,curva=False,color_texto='000000',subrayado=False):
    if (subrayado):
        tipo_subrayado='single'
    else:
        tipo_subrayado=None
    sheet[celda].font = Font(fuente,
                 size=tamaño,
                 bold=negrita,
                 italic=curva,
                 underline=tipo_subrayado,
                 color=color_texto)

def bordear_celdasv1(sheet,celda_ini,celda_final):#A1 , E8
    letra_ini_col = celda_ini[0]#string
    letra_ini_col = ord(letra_ini_col)-64#letra a numero

    letra_final_col = celda_final[0]#string
    letra_final_col = ord(letra_final_col)-64#letra a numero

    letra_ini_fil = int(celda_ini[1])#int
    letra_final_fil = int(celda_final[1])#int

    for fila in range(letra_ini_fil,letra_final_fil):
        for columna in range(letra_ini_col,letra_final_col):
            sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                    right=Side(border_style='thin', color='000000'),
                                                    top=Side(border_style='thin', color='000000'),
                                                    bottom=Side(border_style='thin', color='000000'))

def bordear_celdasv2(sheet,iterable,ultima_fila,centrado=False):#cuando tienes la posicion de las celdas en numeros,ACTUAL EN USO
    for i in range(iterable,ultima_fila+1):
            for j in range(1,6):
                sheet.cell(i,j).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style='thin', color='000000'),
                                                top=Side(border_style='thin', color='000000'),
                                                bottom=Side(border_style='thin', color='000000'))
                if centrado:
                    letra=get_column_letter(j)
                    celda=f'{letra}{i}'
                    sheet[celda].alignment=Alignment(horizontal='center')

def izquierda(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style='thin', color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('izquierda')

def derecha(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style='thin', color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('derecha')

def arriba(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style='thin', color='000000'),
                                                bottom=Side(border_style=None, color='000000'))
    #print('arriba')

def abajo(sheet,fila,columna):
    sheet.cell(fila,columna).border = Border(left=Side(border_style=None, color='000000'),
                                                right=Side(border_style=None, color='000000'),
                                                top=Side(border_style=None, color='000000'),
                                                bottom=Side(border_style='thin', color='000000'))
    #print('abajo')

def bordear_lado(lado,sheet,fila,columna):
    {
        'izquierda': izquierda(sheet,fila,columna),
        'derecha':derecha(sheet,fila,columna),
        'arriba':derecha(sheet,fila,columna),
        'abajo':derecha(sheet,fila,columna)
    }.get(lado)

def getColumnName(n):#vota string
    # inicializa la string de salida como vacía
    result = ''
    while n > 0:

        # encontrar el índice de la siguiente letra y concatenar la letra
        # a la solución
 
        # aquí el índice 0 corresponde a `A`, y 25 corresponde a `Z`
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
    return result[::-1]

def ancho_col(sheet):
    lista_ancho_columnas=[17,18,14,10,17]
    anchos=0
    for columna in range(1,6):
        col_letter = get_column_letter(columna)
        medida=lista_ancho_columnas[anchos]
        sheet.column_dimensions[col_letter].width = medida #NO ESTA EN PIXELES
        anchos=anchos+1

def convertir_pdf(varbuffer,cantidad_propietarios,finca,año,mes): 
    json_rutas_pdf = []
    for i in range(cantidad_propietarios):
        if (varbuffer[i][0]!=1):
            ruta = str(pathlib.Path().absolute())
            a=ruta.replace('\\','/')
            x = a.rfind("/")
            nombre_propietario = varbuffer[i][1]
            ruta_base = get_url_api()
            new_ruta = ruta_base+'/recibos/'+str(mes)+'_propietario_'+nombre_propietario+'.pdf'
            ruta_xls =a[0:x+1] +"fincas/"+finca+"/"+nombre_propietario+'/'+str(año)+'/'+str(mes)+'_propietario_'+nombre_propietario+".xlsx"
            pdf_path = convertir_a_pdf(ruta_xls)
            nombres_completos = varbuffer[i][1]
            estado = varbuffer[i][0]
            diccionario_info = {"cliente":nombres_completos,"estado":estado,"ruta_pdf":new_ruta}
            json_rutas_pdf.append(diccionario_info)
    return json_rutas_pdf

def pdf_versionb2(ruta_excel):
    excel_app = xw.App(visible=False)
    print('Iniciando')
    excel_book = excel_app.books.open(ruta_excel)
    pdf_path = ruta_excel.replace('xlsx','pdf')
    # Save excel workbook to pdf file
    print(f"Saving workbook as '{pdf_path}' ...")
    excel_book.api.ExportAsFixedFormat(0, pdf_path)
    print('conversion exitosa')
    excel_book.close()
    excel_app.quit()
    #convertir_base64 el PDF
    return pdf_path
    
def convertir_a_pdf(ruta_excel):#FALLA SI EL PDF YA EXISTE
    pythoncom.CoInitialize()#nueva adicion
    excel_app = xw.App(visible=False)
    print('Iniciando ...')
    excel_book = excel_app.books.open(ruta_excel)
    try:
        pdf_path = ruta_excel.replace('xlsx','pdf')
        print(f"Saving workbook as '{pdf_path}' ...")
        excel_book.api.ExportAsFixedFormat(0, pdf_path)
        print('conversion exitosa')
        excel_book.close()
        excel_app.quit()
        return pdf_path
    except Exception as e:
        excel_book.close()
        excel_app.quit()
        print(e)
        error = e
        return error

def poner_imagenes(sheet):
    logo_pil = Image.open('../excels/archivos/LOGOVIDEOFINCA_ORIGINAL.png')
    finca_pil = Image.open('../excels/archivos/FINCA.jpg')

    proporcion = 3
    alto = 40
    ancho = int(proporcion * alto)

    logo_pil = logo_pil.resize((ancho,alto+5))#140 80 solo acepta enteros
    logo_pil.save('../excels/archivos/LOGOVIDEOFINCA_MODIFICADO.png')
    finca_pil = finca_pil.resize((ancho+3,alto))
    finca_pil.save('../excels/archivos/FINCA_MODIFICADO.jpg')
    logo = openpyxl.drawing.image.Image('../excels/archivos/LOGOVIDEOFINCA_MODIFICADO.png')
    p2e = pixels_to_EMU
    c2e = cm_to_EMU
    h, w = logo.height, logo.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10)
    column = 0
    coloffset = cellw(0.05)
    row = 1
    rowoffset = cellh(0.05)
    marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
    logo.anchor = OneCellAnchor(_from=marker, ext=size)
    sheet.add_image(logo)

    finca = openpyxl.drawing.image.Image('../excels/archivos/FINCA_MODIFICADO.jpg')
    p2e = pixels_to_EMU
    c2e = cm_to_EMU
    h, w = finca.height, finca.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10)
    column = 4
    coloffset = cellw(0.05)
    row = 1
    rowoffset = cellh(0.05)
    marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
    finca.anchor = OneCellAnchor(_from=marker, ext=size)
    sheet.add_image(finca) 

def ancho_columnas_parametros(sheet):
    lista_ancho_columnas=[17,20,15,20,17]
    anchos=0
    for columna in range(1,6):
        col_letter = get_column_letter(columna)
        medida=lista_ancho_columnas[anchos]
        sheet.column_dimensions[col_letter].width = medida #NO ESTA EN PIXELES
        anchos=anchos+1

def convertir_base64(ruta):# ruta: C:/Users/DELL/Desktop/excel.xlsx
    with open(ruta, 'rb') as binary_file:
        binary_file_data = binary_file.read()
        base64_encoded_data = base64.b64encode(binary_file_data)
        base64_message = base64_encoded_data.decode('utf-8')
        #print(base64_message)
    return base64_message

def decodificar_base64(codificado,nombre_archivo,extension):
    base64_img_bytes = codificado.encode('utf-8')
    with open(f'C:/Users/DELL/Desktop/{nombre_archivo}_decodificado.{extension}', 'wb') as file_to_save:
        decoded_image_data = base64.decodebytes(base64_img_bytes)
        file_to_save.write(decoded_image_data)

def borrar_temporal():
    #eliminando la carpeta fincas
    ruta = str(pathlib.Path().absolute())
    a=ruta.replace('\\','/')
    x = a.rfind("/")
    ruta_temp = a[0:x+1]+"fincas" #Segun la ruta a guardar
    shutil.rmtree(ruta_temp)
    #creandola de vuelta
    path = pathlib.Path(ruta_temp)
    path.mkdir(parents=True)

def get_url_api():
    url = 'http://192.168.195.8:4000'
    return url

# fincas
# --nombrepropietario
#   -- año
#       -- recibo_mes_nombrepropietario

def guardar_ruta_excel(nombre_excel,finca,propietario,año,mes):
    ruta_archivos = 'fincas/'+finca+'/'+propietario+'/'+str(año) #falta /mes_propietario.xlsx
    ruta = str(pathlib.Path().absolute())
    a=ruta.replace('\\','/')
    x = a.rfind("/")
    ruta_carpetas = pathlib.Path(a[0:x+1]+ruta_archivos)
    ruta_carpetas.mkdir(parents=True)
    ruta_excel_real = a[0:x+1]+ruta_archivos+'/'+str(mes)+'_'+nombre_excel
    return ruta_excel_real

def agregar_fecha():
    now = str(datetime.now())
    now = now[:19]
    return now

def generar_id():
    id = str(uuid.uuid4())
    return id

def validar_id_departamento(departamento,base_datos,estado,Nombre_y_Apellidos):
    validacion = False
    engine = conexion(base_datos)
    metadata = MetaData()

    # Primero ve si el cliente esta activo
    
    try:
        table1 = Table('propietarios', metadata, autoload=True, autoload_with=engine)
        query1 = (select([table1]).where((table1.columns.estado == estado)  and (table1.columns.nombres_y_apellidos == Nombre_y_Apellidos)))
        result1 = engine.execute(query1).fetchall()
        #print(type(result1)) #es una lista
        if result1 != '':
            #se incluye el for en caso el propietarios tenga mas fincas/estac/deposi alquilados y otros usados por el
            for row1 in result1:
                idpropiedad = row1.idpropiedad
                #bsucamos el idpropiedad en la tabla propiedad
                table2 = Table('propiedad', metadata, autoload=True, autload=True, autoload_with=engine)
                query2 = (select([table2]).where((table2.columns.idpropiedad == idpropiedad)))
                result2 = engine.execute(query2).fetchall()
                for row2 in result2:
                    #buscamos por for en caso el propietario tenga mas de una combinacion de bienes (fincas/estac/deposi)
                    numero_departamento = row2.numero_departamento
                    if (numero_departamento == departamento):
                        validacion = False
                    else: validacion = True
        return validacion
    except Exception as e:
        print('Error → ', str(e))
        return validacion

def validar_id_estacionamiento(num_estacionamiento,Finca,estado):
    validacion = False
    try:
        respuesta = conexion('propietarios').find({"$and": [
            {"Estacionamientos.Numero_Estacionamiento": f'{num_estacionamiento}'}, 
            {"Finca": f'{Finca}'},
            {"estado": f'{estado}'}
            ]})
        response = json_util.dumps(respuesta)#es un string []
        consulta = json_util.loads(response)#diccionario
        if len(consulta)>0:
            return validacion
        else:
            validacion = True
            return validacion
    except Exception as e:
        print('Error → ', str(e))
        return validacion
def validar_id_deposito(num_deposito,Finca,estado):
    validacion = False
    try:
        respuesta = conexion('propietarios').find({"$and": [
            {"Numero_deposito": f'{num_deposito}'}, 
            {"Finca": f'{Finca}'},
            {"estado": f'{estado}'}
            ]})
        response = json_util.dumps(respuesta)#es un string []
        consulta = json_util.loads(response)#diccionario
        if len(consulta)>0:
            return validacion
        else:
            validacion = True
            return validacion
    except Exception as e:
        print('Error → ', str(e))
        return validacion

def validar_usuario(usuario,operacion='login'):#login o para registrar
    validacion = False
    try:
        respuesta = conexionv2('administradores','lista').find(
            {"user": f'{usuario}'})
        response = json_util.dumps(respuesta)#es un string []
        consulta = json_util.loads(response)#diccionario
        if len(consulta)>0:
            if operacion == 'login':
                validacion = True
                return validacion,consulta
            else:
                return validacion,consulta
        else:
            if operacion == 'login':
                return validacion,consulta
            else:
                validacion = True
                return validacion,consulta
    except Exception as e:
        print('Error → ', str(e))
        return validacion,consulta

def nombre_mes(mes,año):
    m = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Setiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre"
        }
    try:
        periodo = m[mes]+' '+str(año)
        return periodo
    except Exception as e:
        print('ERROR >>>',e)

def contar_porc_participacion(Finca,estado):
    sumatoria = 0
    try:
        respuesta = conexion('propietarios').find({"$and": [
            {"Finca": f'{Finca}'},
            {"estado": f'{estado}'}
            ]})
        response = json_util.dumps(respuesta)#es un string []
        consulta = json_util.loads(response)#diccionario
        long = len(consulta) #cuenta a todos los propietarios
        for i in range(long):
            var = consulta[i]["Departamentos"][0]["Porcentaje_Participacion"]
            sumatoria = sumatoria + float(var) #es un float
        conexion('finca').update_one(
                    {'_id': Finca}, {'$set': {
                                            "Total_porc_participacion": sumatoria}})                                   
    except Exception as e:
        print('ERROR EN ACTUALIZAR PORC DE PARTICIPACION>>> ',e)

def encriptar_contra():
    password = b"123"#tipo bytes
    admin = b"admin1"
    # Hash a password for the first time, with a randomly-generated salt
    hashed = bcrypt.hashpw(password, bcrypt.gensalt())#tipo bytes
    hashed_admin = bcrypt.hashpw(admin, bcrypt.gensalt())
    # Check that an unhashed password matches one that has previously been
    # hashed
    #wrong_password = b'wrong'
    if bcrypt.checkpw(password, hashed):
        print("It Matches!")
    else:
        print("It Does not Match :(")
    texto = byte_a_string(hashed)
    texto_admin = byte_a_string(hashed_admin)
    texto_byte = string_a_byte(texto)
    print(texto_byte)
    usuario_prueba = "user_prueba"
    """db = conexionv2('administradores','lista')
    db.insert_one(
        {
            "user":usuario_prueba,"password": texto,"bd":texto_admin
        }
    )"""

def byte_a_string(palabra):#en byte
    palabra = palabra.decode("utf-8")
    return palabra

def string_a_byte(palabra):#en string
    palabra = str.encode(palabra)
    return palabra

def encriptar(password):#en encriptar
    hashed = bcrypt.hashpw(password, bcrypt.gensalt())
    print(type(hashed))
    return hashed

#PRUEBAS        
def prueba():
    #PONER LA FUNCION A PROBAR Y DESCOMENTAR LA ULTIMA LINEA
    response,json =validar_usuario('user_prueba')
    print(json)
    print('\n')

#prueba()